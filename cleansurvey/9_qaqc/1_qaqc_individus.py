"""
cleansurvey/9_qaqc/1_qaqc_individus.py
======================================
Rapport QAQC Table Individus (Version Intégrée)

Produit :
  - data/output_qaqc/qaqc_individus.xlsx (Rapport Excel multi-onglets)
  - data/output_qaqc/qaqc_individus.html (Rapport HTML)
  - data/output_qaqc/qaqc_individus.md   (Rapport Markdown)
"""

import sys
import os
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl

# Configurer le chemin racine (remonter de 2 niveaux depuis cleansurvey/9_qaqc)
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
# Ajouter le dossier contenant le script au path pour pouvoir importer qaqc_utils directement
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from cleansurvey.config import OUTPUT_DIR, QAQC_DIR
from cleansurvey.utils import section_header

# Import des utilitaires QAQC communs
from qaqc_utils import (
    C_VERT_FONCE, C_VERT_MOYEN, C_VERT_CLAIR, C_VERT_PALE, C_TEXTE, C_BLANC,
    dist_variable, stats_manquants,
    style_cell, write_title, write_headers, write_data, set_widths, onglet_dist,
    df_to_html, carte, section, note_item, bloc_note
)

# ════════════════════════════════════════════════════════════════
# CALCULS DES ESTIMATIONS PRIMAIRES INDIVIDUELLES
# ════════════════════════════════════════════════════════════════

def stats_generales(df):
    males = (df['sexe'] == 'Masculin').sum()
    females = (df['sexe'] == 'Féminin').sum()
    sex_ratio = (males / females * 100) if females > 0 else np.nan
    
    rows = [
        ("Nombre total d'individus", f"{len(df):,}"),
        ("Nombre de variables", str(df.shape[1])),
        ("Effectif Hommes", f"{males:,} ({males/len(df)*100:.1f}%)"),
        ("Effectif Femmes", f"{females:,} ({females/len(df)*100:.1f}%)"),
        ("Sex-ratio (Hommes pour 100 Femmes)", f"{sex_ratio:.2f}" if pd.notna(sex_ratio) else "N/A"),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def stats_age(df):
    if 'age' not in df.columns:
        return pd.DataFrame(columns=["Indicateur", "Valeur"])
    s = df['age']
    mean_age = s.mean()
    median_age = s.median()
    
    # Tranches d'âge
    bins = [0, 5, 12, 15, 18, 25, 65, 120]
    labels = ['0-4 ans', '5-11 ans', '12-14 ans', '15-17 ans', '18-24 ans', '25-64 ans', '65 ans et +']
    groups = pd.cut(s, bins=bins, labels=labels, right=False)
    counts = groups.value_counts(dropna=False)
    pcts = groups.value_counts(normalize=True, dropna=False) * 100
    
    rows = [
        ("Âge moyen", f"{mean_age:.1f} ans"),
        ("Âge médian", f"{median_age:.0f} ans"),
        ("Écart-type", f"{s.std():.1f} ans"),
        ("Minimum", f"{s.min():.0f} ans"),
        ("Maximum", f"{s.max():.0f} ans"),
    ]
    for lbl in labels:
        rows.append((f"Tranche {lbl}", f"{counts.get(lbl, 0):,} ({pcts.get(lbl, 0):.1f}%)"))
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def stats_scolarisation(df):
    if 'scolarisation' not in df.columns or 'age' not in df.columns:
        return pd.DataFrame(columns=["Indicateur", "Valeur"])
    df_prim = df[df['age'].between(6, 11)]
    df_sec = df[df['age'].between(12, 18)]
    
    scol_prim = (df_prim['scolarisation'] == 'oui, fréquente actuellement').mean() * 100 if len(df_prim) > 0 else np.nan
    scol_sec = (df_sec['scolarisation'] == 'oui, fréquente actuellement').mean() * 100 if len(df_sec) > 0 else np.nan
    
    rows = [
        ("Taux de scolarisation active (6-11 ans)", f"{scol_prim:.2f}%" if pd.notna(scol_prim) else "N/A"),
        ("Taux de scolarisation active (12-18 ans)", f"{scol_sec:.2f}%" if pd.notna(scol_sec) else "N/A"),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def stats_alphabetisation(df):
    df_5 = df[df['age'] >= 5]
    if len(df_5) == 0:
        return pd.DataFrame(columns=["Langue", "Taux d'alphabétisation (%)"])
    alpha_cols = [c for c in df.columns if c.startswith('alpha_')]
    alpha_rates = []
    for col in alpha_cols:
        lang_name = col.replace('alpha_', '').upper()
        total_valid = df_5[col].dropna()
        if len(total_valid) > 0:
            is_yes = total_valid.astype(str).str.lower().str.strip().isin(['oui', '1', '1.0']) | \
                     (~total_valid.astype(str).str.lower().str.strip().isin(['non', '0', '0.0', 'nan', 'none']))
            rate = is_yes.mean() * 100
            alpha_rates.append({
                "Langue": lang_name,
                "Taux d'alphabétisation (%)": f"{rate:.2f}%"
            })
    if alpha_rates:
        return pd.DataFrame(alpha_rates).sort_values(by="Taux d'alphabétisation (%)", ascending=False).reset_index(drop=True)
    return pd.DataFrame(columns=["Langue", "Taux d'alphabétisation (%)"])


def stats_revenu_moyen(df):
    if 'statut_emploi' not in df.columns or 'revenu_emploi_estime' not in df.columns or 'age' not in df.columns:
        return pd.DataFrame(columns=["Secteur Institutionnel", "Revenu Moyen Estimé (FCFA)"])
    df_15 = df[df['age'] >= 15]
    df_occ = df_15[df_15['statut_emploi'].isin(['Occupé', '1', '1.0'])]
    if len(df_occ) > 0 and 'secteur_instit' in df.columns:
        rev_secteur = df_occ.groupby('secteur_instit')['revenu_emploi_estime'].mean().round(0).reset_index()
        rev_secteur.columns = ["Secteur Institutionnel", "Revenu Moyen Estimé (FCFA)"]
        rev_secteur["Revenu Moyen Estimé (FCFA)"] = rev_secteur["Revenu Moyen Estimé (FCFA)"].apply(lambda x: f"{x:,.0f} FCFA")
        return rev_secteur
    return pd.DataFrame(columns=["Secteur Institutionnel", "Revenu Moyen Estimé (FCFA)"])


def stats_handicap(df):
    handicap_cols = ['handicap_vision', 'handicap_audition', 'handicap_moteur', 'handicap_cognitif', 'handicap_soins', 'handicap_communication']
    handicap_rates = []
    for col in handicap_cols:
        if col in df.columns:
            rate = (df[col] == 1).mean() * 100
            handicap_rates.append({
                "Type de Limitation (Handicap)": col.replace('handicap_', '').capitalize(),
                "Taux de prévalence (%)": f"{rate:.2f}%"
            })
    if handicap_rates:
        return pd.DataFrame(handicap_rates).sort_values(by="Taux de prévalence (%)", ascending=False).reset_index(drop=True)
    return pd.DataFrame(columns=["Type de Limitation (Handicap)", "Taux de prévalence (%)"])


# ════════════════════════════════════════════════════════════════
# RAPPORT EXCEL
# ════════════════════════════════════════════════════════════════

def onglet_resume(wb, df):
    ws = wb.active
    ws.title = "Résumé général"
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=1,
                value="RAPPORT QAQC - TABLE INDIVIDUS - RGPH-5 Sénégal")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=15, align="center")
    ws.merge_cells("A1:B1")
    row += 2

    row = write_title(ws, row, "STATISTIQUES GÉNÉRALES")
    row = write_headers(ws, row, ["Indicateur", "Valeur"])
    row = write_data(ws, row, stats_generales(df))
    row += 1

    if "age" in df.columns:
        row = write_title(ws, row, "CARACTÉRISTIQUES D'ÂGE")
        row = write_headers(ws, row, ["Indicateur", "Valeur"])
        row = write_data(ws, row, stats_age(df))
        row += 1

    if "scolarisation" in df.columns:
        row = write_title(ws, row, "TAUX DE SCOLARISATION")
        row = write_headers(ws, row, ["Indicateur", "Valeur"])
        row = write_data(ws, row, stats_scolarisation(df))

    set_widths(ws, [45, 35])


def onglet_qualite(wb, df, df_log):
    ws = wb.create_sheet("Qualité des données")
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1,
                value="QUALITÉ DES DONNÉES - VALEURS MANQUANTES")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=13, align="center")
    ws.merge_cells(f"A{row}:D{row}")
    row += 2

    manq = stats_manquants(df)
    row = write_headers(ws, row, ["Variable", "Nb manquants", "Taux (%)", "Statut"])

    for i, r in enumerate(manq.itertuples(index=False)):
        ws.row_dimensions[row].height = 18
        bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
        for c_i, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c_i, value=val)
            if c_i == 4:
                if val == "Complet":
                    style_cell(cell, bold=True, bg="C6F6D5", fg="1A4731",
                               size=10, align="center", border=True)
                elif val == "Faible < 5%":
                    style_cell(cell, bg="FEFCBF", fg="744210",
                               size=10, align="center", border=True)
                else:
                    style_cell(cell, bg="FED7D7", fg="742A2A",
                               size=10, align="center", border=True)
            else:
                style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
        row += 1

    if df_log is not None and len(df_log) > 0:
        row += 2
        ws.row_dimensions[row].height = 28
        c = ws.cell(row=row, column=1, value="LOG DES CORRECTIONS EFFECTUÉES")
        style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=12, align="center")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        row = write_headers(ws, row, ["Contrôle", "Nb cas", "Action", ""])
        df_log_disp = df_log[["controle", "nb_cas", "action"]].copy()
        df_log_disp[""] = ""
        row = write_data(ws, row, df_log_disp)

    set_widths(ws, [42, 12, 55, 5])


def onglet_notes(wb):
    ws = wb.create_sheet("Notes méthodologiques")
    ws.sheet_view.showGridLines = False

    notes = [
        ("TRAITEMENT DES VALEURS ABERRANTES & IMPUTATIONS", [
            ("Âge (age)",
             "Bornes définies : [0, 120] ans. Imputation par la médiane. Toute valeur hors de ces bornes est corrigée."),
            ("Variables catégorielles",
             "Imputation par le mode ou affectation à 'na'/'Unknown' selon les stratégies de config.py."),
        ]),
        ("CONTRÔLES DE COHÉRENCE LOGIQUE", [
            ("Mariage précoce",
             "Si l'âge < 12 ans et situation matrimoniale non célibataire, celle-ci est corrigée en 'Célibataire'."),
            ("Éducation précoce",
             "Si l'âge < 3 ans et scolarisé, scolarisation est corrigée en 'non, n'a jamais fréquenté'."),
            ("Sexe et Épouse",
             "Si sexe différent de Féminin et lien de parenté est Épouse du CM, le sexe est corrigé en 'Féminin'."),
            ("Travail des enfants",
             "Si l'âge < 10 ans et statut d'emploi actif (occupé/chômeur), le statut est corrigé en 'Autres inactifs'."),
        ]),
    ]

    row = 1
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=1,
                value="NOTES MÉTHODOLOGIQUES - DÉCISIONS DE TRAITEMENT")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=14, align="center")
    ws.merge_cells(f"A{row}:B{row}")
    row += 2

    for section_titre, items in notes:
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=1, value=section_titre)
        style_cell(c, bold=True, bg=C_VERT_MOYEN, fg=C_BLANC,
                   size=11, align="left")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        row = write_headers(ws, row, ["Point", "Explication"])

        for i, (point, explication) in enumerate(items):
            ws.row_dimensions[row].height = 45
            bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
            for c_i, val in enumerate([point, explication], 1):
                cell = ws.cell(row=row, column=c_i, value=val)
                style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
            row += 1
        row += 1

    set_widths(ws, [35, 80])


def generer_rapport_excel(df, df_log, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    onglet_resume(wb, df)
    onglet_dist(wb, df, "sexe", "Sexe", "RÉPARTITION PAR SEXE")
    onglet_dist(wb, df, "lien_parente", "Lien parenté", "RÉPARTITION PAR LIEN DE PARENTÉ AVEC LE CM")
    onglet_dist(wb, df, "situation_matrimoniale", "Matrimonial", "SITUATION MATRIMONIALE (15 ans et +)")
    onglet_dist(wb, df, "niveau_etudes", "Niveau études", "RÉPARTITION PAR NIVEAU D'ÉTUDES")
    onglet_dist(wb, df, "region", "Régions", "DISTRIBUTION DES INDIVIDUS PAR RÉGION")
    
    milieu_col = "milieu" if "milieu" in df.columns else ("milieu_residence" if "milieu_residence" in df.columns else None)
    if milieu_col:
        onglet_dist(wb, df, milieu_col, "Milieu", "DISTRIBUTION DES INDIVIDUS PAR MILIEU DE RÉSIDENCE")
        
    onglet_dist(wb, df, "statut_emploi", "Statut emploi", "RÉPARTITION PAR STATUT D'EMPLOI")
    onglet_dist(wb, df, "secteur_instit", "Secteur activité", "SECTEUR D'ACTIVITÉ (15 ans et +)")
    onglet_dist(wb, df, "situation_residence", "Résidence", "SITUATION DE RÉSIDENCE")
    
    onglet_qualite(wb, df, df_log)
    onglet_notes(wb)

    wb.save(output_path)
    print(f"  Rapport Excel : {output_path}")


# ════════════════════════════════════════════════════════════════
# RAPPORT HTML
# ════════════════════════════════════════════════════════════════

def generer_rapport_html(df, df_log, output_path):
    milieu_col = "milieu" if "milieu" in df.columns else ("milieu_residence" if "milieu_residence" in df.columns else None)
    
    milieu_urb_pct = f"{(df[milieu_col]=='Urbain').mean()*100:.1f}%" if milieu_col else "N/A"
    milieu_rur_pct = f"{(df[milieu_col]=='Rural').mean()*100:.1f}%" if milieu_col else "N/A"
    
    sexe_m_pct = f"{(df['sexe']=='Masculin').mean()*100:.1f}%" if "sexe" in df.columns else "N/A"
    sexe_f_pct = f"{(df['sexe']=='Féminin').mean()*100:.1f}%" if "sexe" in df.columns else "N/A"
    
    age_med = f"{df['age'].median():.0f} ans" if "age" in df.columns else "N/A"
    
    scol_prim_val = "N/A"
    if 'scolarisation' in df.columns and 'age' in df.columns:
        df_prim = df[df['age'].between(6, 11)]
        if len(df_prim) > 0:
            scol_prim_val = f"{(df_prim['scolarisation'] == 'oui, fréquente actuellement').mean() * 100:.1f}%"

    handicap_val = "N/A"
    handicap_cols = ['handicap_vision', 'handicap_audition', 'handicap_moteur', 'handicap_cognitif', 'handicap_soins', 'handicap_communication']
    avail_cols = [c for c in handicap_cols if c in df.columns]
    if avail_cols:
        df['any_handicap'] = (df[avail_cols] == 1).any(axis=1).astype(int)
        handicap_val = f"{df['any_handicap'].mean() * 100:.1f}%"

    # Cartes
    cartes = f"""
    <div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:8px;">
         {carte("Individus traités",      f"{len(df):,}",                                       "#1A4731")}
         {carte("Milieu urbain",         milieu_urb_pct,                                       "#276749")}
         {carte("Milieu rural",          milieu_rur_pct,                                       "#276749")}
         {carte("Hommes",                sexe_m_pct,                                           "#2F855A")}
         {carte("Femmes",                sexe_f_pct,                                           "#2F855A")}
         {carte("Âge médian",            age_med,                                              "#38A169")}
         {carte("Scolarisation (6-11)",  scol_prim_val,                                        "#48BB78")}
         {carte("Prévalence Handicap",   handicap_val,                                         "#48BB78")}
    </div>"""

    # Notes méthodologiques
    notes_html = "".join([
        bloc_note("Traitement des valeurs aberrantes", [
            ("Âge", "Bornes définies : [0, 120] ans. Imputation par la médiane en cas de valeur hors-bornes."),
        ]),
        bloc_note("Contrôles de cohérence", [
            ("Mariage précoce", "Si l'âge < 12 ans et situation matrimoniale non célibataire, celle-ci est corrigée en 'Célibataire'."),
            ("Éducation précoce", "Si l'âge < 3 ans et scolarisé, scolarisation est corrigée en 'non, n'a jamais fréquenté'."),
            ("Sexe et Épouse", "Si sexe différent de Féminin et lien de parenté est Épouse du CM, le sexe est corrigé en 'Féminin'."),
            ("Travail des enfants", "Si l'âge < 10 ans et statut d'emploi actif (occupé/chômeur), le statut est corrigé en 'Autres inactifs'."),
        ]),
    ])

    # Grille de tables dynamiques
    sections_dynamiques = []
    
    # Age et Démographie générale
    dem_html = ""
    if "age" in df.columns:
        dem_html += f'<div class="card">{section("Structure par Âge", df_to_html(stats_age(df)))}</div>'
    if "sexe" in df.columns:
        dem_html += f'<div class="card">{section("Distribution par Sexe", df_to_html(dist_variable(df, "sexe")))}</div>'
    if dem_html:
        sections_dynamiques.append(f'<div class="grid-2">{dem_html}</div><br>')
        
    # Parenté et situation matrimoniale
    mat_html = ""
    if "lien_parente" in df.columns:
        mat_html += f'<div class="card">{section("Lien de parenté avec le CM", df_to_html(dist_variable(df, "lien_parente")))}</div>'
    if "situation_matrimoniale" in df.columns:
        df_15 = df[df['age'] >= 15] if 'age' in df.columns else df
        mat_html += f'<div class="card">{section("Situation matrimoniale (15 ans et +)", df_to_html(dist_variable(df_15, "situation_matrimoniale")))}</div>'
    if mat_html:
        sections_dynamiques.append(f'<div class="grid-2">{mat_html}</div><br>')

    # Scolarisation et Alphabétisation
    scol_html = ""
    if "scolarisation" in df.columns:
        scol_html += f'<div class="card">{section("Taux de scolarisation active", df_to_html(stats_scolarisation(df)))}</div>'
    alpha_df = stats_alphabetisation(df)
    if not alpha_df.empty:
        scol_html += f'<div class="card">{section("Alphabétisation par langue (5 ans et +)", df_to_html(alpha_df))}</div>'
    if scol_html:
        sections_dynamiques.append(f'<div class="grid-2">{scol_html}</div><br>')

    # Niveau d'études et Régions
    reg_html = ""
    if "niveau_etudes" in df.columns:
        lbl_etudes = "Niveau d'études"
        reg_html += f'<div class="card">{section(lbl_etudes, df_to_html(dist_variable(df, "niveau_etudes")))}</div>'
    if "region" in df.columns:
        reg_html += f'<div class="card">{section("Distribution par région", df_to_html(dist_variable(df, "region")))}</div>'
    if reg_html:
        sections_dynamiques.append(f'<div class="grid-2">{reg_html}</div><br>')
        
    # Emploi et Secteurs
    emp_html = ""
    df_15 = df[df['age'] >= 15] if 'age' in df.columns else df
    if "statut_emploi" in df.columns:
        lbl_statut = "Statut d'emploi (15 ans et +)"
        emp_html += f'<div class="card">{section(lbl_statut, df_to_html(dist_variable(df_15, "statut_emploi")))}</div>'
    if "secteur_instit" in df.columns:
        emp_html += f'<div class="card">{section("Secteur institutionnel (15 ans et +)", df_to_html(dist_variable(df_15, "secteur_instit")))}</div>'
    if emp_html:
        sections_dynamiques.append(f'<div class="grid-2">{emp_html}</div><br>')

    # Revenu par secteur et professions top 10
    rev_html = ""
    rev_df = stats_revenu_moyen(df)
    if not rev_df.empty:
        rev_html += f'<div class="card">{section("Revenu moyen estimé par secteur", df_to_html(rev_df))}</div>'
    if "profession" in df.columns:
        df_prof = dist_variable(df_15, "profession").head(10)
        rev_html += f'<div class="card">{section("Top 10 Professions", df_to_html(df_prof))}</div>'
    if rev_html:
        sections_dynamiques.append(f'<div class="grid-2">{rev_html}</div><br>')

    # Handicap
    handi_df = stats_handicap(df)
    if not handi_df.empty:
        sections_dynamiques.append(f'<div class="card">{section("Limitations et handicap (Washington Group)", df_to_html(handi_df))}</div><br>')

    dinamic_content = "\n".join(sections_dynamiques)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>QAQC - Table Individus - RGPH-5</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
      background: #f7faf7;
      color: #2d3748;
      line-height: 1.6;
  }}
  .header {{
      background: linear-gradient(135deg, #1A4731 0%, #276749 60%, #38A169 100%);
      color: white;
      padding: 44px 52px;
  }}
  .header h1 {{
      font-size: 26px;
      font-weight: 800;
      letter-spacing: -0.02em;
      margin-bottom: 8px;
  }}
  .header p {{ font-size: 14px; opacity: 0.85; margin-top: 4px; }}
  .badge {{
      display: inline-block;
      background: rgba(255,255,255,0.18);
      border: 1px solid rgba(255,255,255,0.35);
      border-radius: 20px;
      padding: 5px 16px;
      font-size: 12px;
      font-weight: 700;
      margin-top: 14px;
      letter-spacing: 0.05em;
  }}
  .container {{
      max-width: 1120px;
      margin: 0 auto;
      padding: 44px 36px;
  }}
  .grid-2 {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
  }}
  .card {{
      background: white;
      border-radius: 12px;
      padding: 26px;
      box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }}
  .footer {{
      text-align: center;
      padding: 22px;
      font-size: 12px;
      color: #718096;
      border-top: 1px solid #c6f6d5;
      margin-top: 40px;
  }}
</style>
</head>
<body>

<div class="header">
    <h1>Rapport QAQC - Table Individus</h1>
    <p>5ème Recensement Général de la Population et de l'Habitat - Sénégal</p>
    <p>Échantillon 1/10 de la base nationale</p>
    <span class="badge">Groupe 2 - Pipeline de traitement</span>
</div>

<div class="container">

    {section("Vue d'ensemble", cartes)}

    {dinamic_content}

    <div class="card">
        {section("Qualité des données - Valeurs manquantes", df_to_html(stats_manquants(df)))}
    </div><br>

    <div class="card">
        {section("Notes méthodologiques - Décisions de traitement", notes_html)}
    </div>

</div>

<div class="footer">
    Généré automatiquement par le pipeline RGPH-5 - Groupe 2
</div>

</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")
    print(f"  Rapport HTML  : {output_path}")


# ════════════════════════════════════════════════════════════════
# RAPPORT MARKDOWN (ORIGINAL)
# ════════════════════════════════════════════════════════════════

def generer_rapport_markdown(df, output_path):
    total_obs = len(df)
    report = []
    report.append("# Rapport d'Assurance Qualité (QAQC) - Recensement RGPH-5 (Groupe 2 - Individus)")
    report.append(f"Ce rapport contient les estimations descriptives primaires après traitement des données d'individus.")
    report.append(f"**Nombre total d'observations traitées** : {total_obs:,} individus.\n")
    report.append("---")
    
    # 1. ANALYSE DES VALEURS MANQUANTES
    report.append("## 1. Analyse des valeurs manquantes (Taux de vide)")
    missing_df = stats_manquants(df)[["Variable", "Nb manquants", "Taux (%)", "Statut"]]
    missing_df.columns = ["Variable", "Nombre de manquants", "Taux de valeurs manquantes (%)", "Statut"]
    report.append(missing_df.to_markdown(index=False))
    report.append("\n")
    
    # 2. CARACTÉRISTIQUES DÉMOGRAPHIQUES
    report.append("## 2. Caractéristiques Démographiques")
    
    # Sexe
    report.append("### Répartition par Sexe (Genre)")
    if 'sexe' in df.columns:
        sexe_df = dist_variable(df, 'sexe', label_name="Sexe")
        sexe_df.columns = ["Sexe", "Effectif", "Pourcentage (%)"]
        report.append(sexe_df.to_markdown(index=False))
        
        males = (df['sexe'] == 'Masculin').sum()
        females = (df['sexe'] == 'Féminin').sum()
        if females > 0:
            sex_ratio = (males / females) * 100
            report.append(f"\n* **Sex-ratio** : {sex_ratio:.2f} hommes pour 100 femmes.\n")
    
    # Âge
    report.append("### Structure par Âge")
    if 'age' in df.columns:
        mean_age = df['age'].mean()
        median_age = df['age'].median()
        
        bins = [0, 5, 12, 15, 18, 25, 65, 120]
        labels = ['0-4 ans', '5-11 ans', '12-14 ans', '15-17 ans', '18-24 ans', '25-64 ans', '65 ans et +']
        df['age_group'] = pd.cut(df['age'], bins=bins, labels=labels, right=False)
        
        age_group_df = dist_variable(df, 'age_group', label_name="Tranche d'âge")
        age_group_df.columns = ["Tranche d'âge", "Effectif", "Pourcentage (%)"]
        
        report.append(f"- **Âge moyen** : {mean_age:.1f} ans")
        report.append(f"- **Âge médian** : {median_age:.1f} ans\n")
        report.append(age_group_df.to_markdown(index=False))
        report.append("\n")
        
    # Situation matrimoniale
    report.append("### Situation Matrimoniale (15 ans et +)")
    if 'situation_matrimoniale' in df.columns:
        df_15 = df[df['age'] >= 15] if 'age' in df.columns else df
        mat_df = dist_variable(df_15, 'situation_matrimoniale', label_name="Situation matrimoniale")
        mat_df.columns = ["Situation matrimoniale", "Effectif", "Pourcentage (%)"]
        report.append(mat_df.to_markdown(index=False))
        report.append("\n")
        
    # Lien de parenté
    report.append("### Lien de Parenté avec le CM")
    if 'lien_parente' in df.columns:
        lp_df = dist_variable(df, 'lien_parente', label_name="Lien de parenté")
        lp_df.columns = ["Lien de parenté", "Effectif", "Pourcentage (%)"]
        report.append(lp_df.to_markdown(index=False))
        report.append("\n")
        
    # Niveau d'études
    report.append("### Niveau d'Études")
    if 'niveau_etudes' in df.columns:
        ne_df = dist_variable(df, 'niveau_etudes', label_name="Niveau d'études")
        ne_df.columns = ["Niveau d'études", "Effectif", "Pourcentage (%)"]
        report.append(ne_df.to_markdown(index=False))
        report.append("\n")

    # 3. CARACTÉRISTIQUES GÉOGRAPHIQUES
    report.append("## 3. Répartition Géographique")
    
    if 'region' in df.columns:
        report.append("### Répartition par Région")
        reg_df = dist_variable(df, 'region', label_name="Région")
        reg_df.columns = ["Région", "Effectif", "Pourcentage (%)"]
        report.append(reg_df.to_markdown(index=False))
        report.append("\n")
        
    if 'milieu_residence' in df.columns:
        report.append("### Répartition par Milieu de Résidence")
        mil_df = dist_variable(df, 'milieu_residence', label_name="Milieu de résidence")
        mil_df.columns = ["Milieu de résidence", "Effectif", "Pourcentage (%)"]
        report.append(mil_df.to_markdown(index=False))
        report.append("\n")
        
    # 4. CARACTÉRISTIQUES ÉDUCATIVES
    report.append("## 4. Caractéristiques Éducatives")
    
    if 'scolarisation' in df.columns and 'age' in df.columns:
        report.append("### Scolarisation")
        df_prim = df[df['age'].between(6, 11)]
        df_sec = df[df['age'].between(12, 18)]
        
        report.append("#### Taux de scolarisation chez les enfants de 6-11 ans (âge du primaire)")
        if len(df_prim) > 0:
            scol_prim_pct = (df_prim['scolarisation'] == 'oui, fréquente actuellement').mean() * 100
            report.append(f"* **Taux d'inscription scolaire active (6-11 ans)** : {scol_prim_pct:.2f}%\n")
            
        report.append("#### Taux de scolarisation chez les adolescents de 12-18 ans (âge du moyen/secondaire)")
        if len(df_sec) > 0:
            scol_sec_pct = (df_sec['scolarisation'] == 'oui, fréquente actuellement').mean() * 100
            report.append(f"* **Taux d'inscription scolaire active (12-18 ans)** : {scol_sec_pct:.2f}%\n")
            
    # Alphabétisation
    report.append("### Alphabétisation par langue (Population de 5 ans et plus)")
    alpha_df = stats_alphabetisation(df)
    if not alpha_df.empty:
        report.append(alpha_df.to_markdown(index=False))
        report.append("\n")
        
    # 5. CARACTÉRISTIQUES DE L'EMPLOI
    report.append("## 5. Caractéristiques de l'Emploi (Population de 15 ans et +)")
    df_15 = df[df['age'] >= 15] if 'age' in df.columns else df
    if len(df_15) > 0:
        if 'statut_emploi' in df.columns:
            report.append("### Répartition par Statut d'Emploi")
            df_se = dist_variable(df_15, 'statut_emploi', label_name="Statut")
            df_se.columns = ["Statut", "Effectif", "Pourcentage (%)"]
            report.append(df_se.to_markdown(index=False))
            report.append("\n")
            
        if 'secteur_instit' in df.columns:
            report.append("### Répartition par Secteur d'Activité (Secteur Institutionnel)")
            df_si = dist_variable(df_15, 'secteur_instit', label_name="Secteur")
            df_si.columns = ["Secteur", "Effectif", "Pourcentage (%)"]
            report.append(df_si.to_markdown(index=False))
            report.append("\n")
            
        if 'profession' in df.columns:
            report.append("### Top 10 des Professions les plus représentées")
            df_prof = dist_variable(df_15, 'profession', label_name="Profession").head(10)
            df_prof.columns = ["Profession", "Effectif", "Pourcentage (%)"]
            report.append(df_prof.to_markdown(index=False))
            report.append("\n")
            
        rev_df = stats_revenu_moyen(df)
        if not rev_df.empty:
            report.append("### Estimation du Revenu d'Emploi Moyen (FCFA)")
            report.append(rev_df.to_markdown(index=False))
            report.append("\n")

    # 6. ANALYSE DU HANDICAP
    report.append("## 6. Analyse du Handicap (Washington Group)")
    handi_df = stats_handicap(df)
    if not handi_df.empty:
        report.append(handi_df.to_markdown(index=False))
        report.append("\n")
        
        handicap_cols = ['handicap_vision', 'handicap_audition', 'handicap_moteur', 'handicap_cognitif', 'handicap_soins', 'handicap_communication']
        avail_cols = [c for c in handicap_cols if c in df.columns]
        if avail_cols:
            df['any_handicap'] = (df[avail_cols] == 1).any(axis=1).astype(int)
            global_prev = df['any_handicap'].mean() * 100
            report.append(f"* **Taux de prévalence globale (au moins une limitation sévère)** : {global_prev:.2f}%\n")

    # 7. ANALYSE DE LA MIGRATION
    report.append("## 7. Analyse de la Migration et Résidence")
    if 'situation_residence' in df.columns:
        report.append("### Répartition de la population par Situation de Résidence")
        df_sr = dist_variable(df, 'situation_residence', label_name="Situation de résidence")
        df_sr.columns = ["Situation de résidence", "Effectif", "Pourcentage (%)"]
        report.append(df_sr.to_markdown(index=False))
        report.append("\n")
        
    if 'region_residence_1an' in df.columns and 'region' in df.columns:
        report.append("### Taux de migration interne récente (inter-régionale) depuis 1 an")
        df_mig = df[df['region_residence_1an'].notna() & (df['region_residence_1an'] != df['region'])]
        mig_rate = (len(df_mig) / len(df)) * 100
        report.append(f"* **Taux de migration interne récente (inter-régionale)** : {mig_rate:.2f}% de la population.\n")
        
        if len(df_mig) > 0:
            report.append("#### Top 5 des régions de destination des migrants récents")
            dest = df_mig['region'].value_counts().head(5).reset_index()
            dest.columns = ["Région de destination", "Nombre de migrants"]
            report.append(dest.to_markdown(index=False))
            report.append("\n")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(report))
    print(f"  Rapport Markdown : {output_path}")


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════

def main():
    section_header("ÉTAPE 5 - RAPPORT QAQC TABLE INDIVIDUS")

    merged_file = os.path.join(OUTPUT_DIR, "rgph5_merged.csv")
    if not os.path.exists(merged_file):
        merged_file = os.path.join(OUTPUT_DIR, "rgph5_ind_clean.csv")
        
    if not os.path.exists(merged_file):
        raise FileNotFoundError(
            f"Introuvable : {merged_file}\n- Lancer d'abord le nettoyage ou la fusion"
        )
        
    df = pd.read_csv(merged_file)
    print(f"\n  Table individus chargée : {len(df):,} lignes × {df.shape[1]} colonnes")

    df_log = pd.DataFrame(columns=["controle", "nb_cas", "action"])

    qaqc_dir = Path(QAQC_DIR)
    qaqc_dir.mkdir(parents=True, exist_ok=True)

    print("\n[Génération des rapports individus]")
    generer_rapport_excel(df, df_log, qaqc_dir / "qaqc_individus.xlsx")
    generer_rapport_html(df, df_log,  qaqc_dir / "qaqc_individus.html")
    generer_rapport_markdown(df, qaqc_dir / "qaqc_individus.md")

    print(f"\n  Rapports disponibles dans : {qaqc_dir}")


if __name__ == "__main__":
    main()
