"""
cleansurvey/9_qaqc/qaqc_utils.py
================================
Fonctions utilitaires communes pour la génération de rapports QAQC (Excel et HTML/Markdown).
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palette verte homogène
C_VERT_FONCE  = "1A4731"
C_VERT_MOYEN  = "276749"
C_VERT_CLAIR  = "C6F6D5"
C_VERT_PALE   = "F0FFF4"
C_TEXTE       = "1C3829"
C_BLANC       = "FFFFFF"
C_GRIS        = "718096"

def dist_variable(df, col, label_name="Modalité"):
    """Calcule la distribution d'une variable catégorielle."""
    if col not in df.columns:
        return pd.DataFrame(columns=[label_name, "Effectif", "Proportion (%)"])
    dist = df[col].value_counts(dropna=False).reset_index()
    dist.columns = [label_name, "Effectif"]
    dist["Proportion (%)"] = (dist["Effectif"] / len(df) * 100).round(2)
    return dist.sort_values("Effectif", ascending=False).reset_index(drop=True)

def stats_manquants(df):
    """Calcule le nombre et le taux de valeurs manquantes par variable."""
    m = df.isnull().sum().reset_index()
    m.columns = ["Variable", "Nb manquants"]
    m["Taux (%)"] = (m["Nb manquants"] / len(df) * 100).round(2)
    m["Statut"] = m["Taux (%)"].apply(
        lambda x: "Complet" if x == 0
        else ("Faible < 5%" if x < 5
        else ("Modéré 5–20%" if x < 20
        else "Élevé > 20%"))
    )
    return m.sort_values("Taux (%)", ascending=False).reset_index(drop=True)

# ── Fonctions utilitaires Excel ─────────────────────────────────

def style_cell(cell, bold=False, bg=None, fg="000000", size=10, align="left", border=False):
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        s = Side(style="thin", color="D0E8D8")
        cell.border = Border(left=s, right=s, top=s, bottom=s)

def write_title(ws, row, titre, nb_cols=2):
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=titre)
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=12, align="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    return row + 1

def write_headers(ws, row, cols):
    ws.row_dimensions[row].height = 22
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=col)
        style_cell(cell, bold=True, bg=C_VERT_MOYEN, fg=C_BLANC, size=10, align="center", border=True)
    return row + 1

def write_data(ws, row, df):
    for i, r in enumerate(df.itertuples(index=False)):
        ws.row_dimensions[row].height = 18
        bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
        for c, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c, value=val)
            style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
        row += 1
    return row

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def onglet_dist(wb, df, col, nom_onglet, titre):
    if col not in df.columns:
        return
    ws = wb.create_sheet(nom_onglet)
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=titre)
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=13, align="center")
    ws.merge_cells(f"A{row}:C{row}")
    row += 2

    dist = dist_variable(df, col)
    row = write_headers(ws, row, ["Modalité", "Effectif", "Proportion (%)"])
    row = write_data(ws, row, dist)

    ws.row_dimensions[row].height = 20
    for c_i, val in enumerate(["TOTAL", f"{len(df):,}", "100.00"], 1):
        cell = ws.cell(row=row, column=c_i, value=val)
        style_cell(cell, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=10, border=True)
    set_widths(ws, [40, 15, 18])

# ── Fonctions utilitaires HTML ──────────────────────────────────

def df_to_html(df):
    if len(df) == 0:
        return "<p style='color:#718096;'>Aucune donnée disponible</p>"
    rows_html = ""
    for i, row in enumerate(df.itertuples(index=False)):
        bg = "#f0fff4" if i % 2 == 0 else "#ffffff"
        cells = "".join(
            f'<td style="padding:10px 14px;border-bottom:1px solid #c6f6d5;color:#1c3829;">{val}</td>'
            for val in row
        )
        rows_html += f'<tr style="background:{bg};">{cells}</tr>'

    headers = "".join(
        f'<th style="padding:12px 14px;text-align:left;font-weight:600;color:#ffffff;background:#276749;letter-spacing:0.03em;">{col}</th>'
        for col in df.columns
    )
    return f"""
    <table style="width:100%;border-collapse:collapse;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.07);">
         <thead><tr>{headers}</tr></thead>
         <tbody>{rows_html}</tbody>
    </table>"""

def carte(label, valeur, couleur="#1A4731"):
    return f"""
    <div style="background:#ffffff;border-radius:10px;padding:20px 22px;box-shadow:0 2px 6px rgba(0,0,0,0.06);border-left:4px solid {couleur};flex:1;min-width:160px;">
        <div style="font-size:10px;text-transform:uppercase;letter-spacing:0.08em;color:#4a5568;font-weight:700;margin-bottom:6px;">{label}</div>
        <div style="font-size:22px;font-weight:800;color:{couleur};">{valeur}</div>
    </div>"""

def section(titre, contenu):
    return f"""
    <div style="margin-bottom:32px;">
         <h2 style="font-size:13px;font-weight:700;color:#1A4731;text-transform:uppercase;letter-spacing:0.07em;border-bottom:2px solid #276749;padding-bottom:7px;margin-bottom:14px;">{titre}</h2>
         {contenu}
    </div>"""

def note_item(point, explication):
    return f"""
    <div style="display:grid;grid-template-columns:220px 1fr;gap:0;border-bottom:1px solid #c6f6d5;">
        <div style="padding:12px 14px;background:#f0fff4;font-weight:600;color:#1A4731;font-size:13px;border-right:2px solid #276749;">{point}</div>
        <div style="padding:12px 16px;color:#2d3748;font-size:13px;line-height:1.6;">{explication}</div>
    </div>"""

def bloc_note(titre_bloc, items):
    items_html = "".join(note_item(p, e) for p, e in items)
    return f"""
    <div style="margin-bottom:24px;border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.07);">
        <div style="background:#276749;color:#ffffff;padding:10px 16px;font-size:12px;font-weight:700;letter-spacing:0.05em;text-transform:uppercase;">{titre_bloc}</div>
        {items_html}
    </div>"""
