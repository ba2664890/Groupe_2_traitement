"""
cleansurvey/9_qaqc/1_survey_data_qaqc/2_qaqc_menage.py
======================================================
Rapport QAQC Table Ménage (Version Intégrée)

Produit :
  - data/output_qaqc/qaqc_menage.xlsx  (rapport Excel multi-onglets)
  - data/output_qaqc/qaqc_menage.html  (rapport HTML)
"""

import sys
import os
from pathlib import Path

# Remonter de 3 niveaux pour trouver cleansurvey/
sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from cleansurvey.config import OUTPUT_DIR, QAQC_DIR
from cleansurvey.utils import section_header

# ── Palette verte ─────────────────────────────────────────────────
C_VERT_FONCE  = "1A4731"
C_VERT_MOYEN  = "276749"
C_VERT_CLAIR  = "C6F6D5"
C_VERT_PALE   = "F0FFF4"
C_TEXTE       = "1C3829"
C_BLANC       = "FFFFFF"
C_GRIS        = "718096"


# ════════════════════════════════════════════════════════════════
# CALCULS DES ESTIMATIONS PRIMAIRES
# ════════════════════════════════════════════════════════════════

def stats_generales(df):
    milieu_col = "milieu" if "milieu" in df.columns else ("milieu_residence" if "milieu_residence" in df.columns else None)
    
    rows = [
        ("Nombre total de ménages",     f"{len(df):Zs}"),
        ("Nombre de variables",          str(df.shape[1])),
    ]
    
    # Correction of formatting and safely evaluating if column exists
    rows = [
        ("Nombre total de ménages",     f"{len(df):,}"),
        ("Nombre de variables",          str(df.shape[1])),
    ]
    
    if milieu_col:
        rows.extend([
            ("Ménages en milieu urbain",     f"{(df[milieu_col]=='Urbain').sum():,}  ({(df[milieu_col]=='Urbain').mean()*100:.1f}%)"),
            ("Ménages en milieu rural",      f"{(df[milieu_col]=='Rural').sum():,}  ({(df[milieu_col]=='Rural').mean()*100:.1f}%)"),
        ])
        
    if "type_menage" in df.columns:
        rows.extend([
            ("Ménages ordinaires",           f"{(df['type_menage']=='Ordinaire').sum():,}  ({(df['type_menage']=='Ordinaire').mean()*100:.1f}%)"),
            ("Ménages collectifs",           f"{(df['type_menage']=='Collectif').sum():,}  ({(df['type_menage']=='Collectif').mean()*100:.1f}%)"),
        ])
        
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def stats_taille(df):
    if "taille_menage" not in df.columns:
        return pd.DataFrame(columns=["Indicateur", "Valeur"])
        
    s = df["taille_menage"]
    rows = [
        ("Taille moyenne",      f"{s.mean():.2f}"),
        ("Taille médiane",      f"{s.median():.0f}"),
        ("Écart-type",          f"{s.std():.2f}"),
        ("Minimum",             f"{s.min():.0f}"),
        ("Maximum",             f"{s.max():.0f}"),
        ("1–2 membres",         f"{((s>=1)&(s<=2)).sum():,}  ({((s>=1)&(s<=2)).mean()*100:.1f}%)"),
        ("3–5 membres",         f"{((s>=3)&(s<=5)).sum():,}  ({((s>=3)&(s<=5)).mean()*100:.1f}%)"),
        ("6–10 membres",        f"{((s>=6)&(s<=10)).sum():,}  ({((s>=6)&(s<=10)).mean()*100:.1f}%)"),
        ("11–20 membres",       f"{((s>=11)&(s<=20)).sum():,}  ({((s>=11)&(s<=20)).mean()*100:.1f}%)"),
        ("21–50 membres",       f"{((s>=21)&(s<=50)).sum():,}  ({((s>=21)&(s<=50)).mean()*100:.1f}%)"),
        ("Plus de 50 membres",  f"{(s>50).sum():,}  ({(s>50).mean()*100:.1f}%)"),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def stats_age_cm(df):
    if "age_cm" not in df.columns:
        return pd.DataFrame(columns=["Indicateur", "Valeur"])
        
    s = df["age_cm"]
    rows = [
        ("Âge moyen",           f"{s.mean():.1f} ans"),
        ("Âge médian",          f"{s.median():.0f} ans"),
        ("Écart-type",          f"{s.std():.1f} ans"),
        ("Minimum",             f"{s.min():.0f} ans"),
        ("Maximum",             f"{s.max():.0f} ans"),
        ("Moins de 25 ans",     f"{(s<25).sum():,}  ({(s<25).mean()*100:.1f}%)"),
        ("25–34 ans",           f"{((s>=25)&(s<35)).sum():,}  ({((s>=25)&(s<35)).mean()*100:.1f}%)"),
        ("35–44 ans",           f"{((s>=35)&(s<45)).sum():,}  ({((s>=35)&(s<45)).mean()*100:.1f}%)"),
        ("45–54 ans",           f"{((s>=45)&(s<55)).sum():,}  ({((s>=45)&(s<55)).mean()*100:.1f}%)"),
        ("55–64 ans",           f"{((s>=55)&(s<65)).sum():,}  ({((s>=55)&(s<65)).mean()*100:.1f}%)"),
        ("65 ans et plus",      f"{(s>=65).sum():,}  ({(s>=65).mean()*100:.1f}%)"),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def dist_variable(df, col):
    if col not in df.columns:
        return pd.DataFrame(columns=["Modalité", "Effectif", "Proportion (%)"])
    dist = df[col].value_counts().reset_index()
    dist.columns = ["Modalité", "Effectif"]
    dist["Proportion (%)"] = (dist["Effectif"] / len(df) * 100).round(2)
    return dist.sort_values("Effectif", ascending=False).reset_index(drop=True)


def stats_manquants(df):
    m = df.isnull().sum().reset_index()
    m.columns = ["Variable", "Nb manquants"]
    m["Taux (%)"] = (m["Nb manquants"] / len(df) * 100).round(2)
    m["Statut"] = m["Taux (%)"].apply(
        lambda x: "Complet" if x == 0
        else ("Faible < 5%" if x < 5
        else ("Modéré 5–20%" if x < 20
        else "Élevé > 20%"))
    )
    return m.sort_values("Taux (%)", ascending=False).reset_index(drop=True)


# ════════════════════════════════════════════════════════════════
# RAPPORT EXCEL
# ════════════════════════════════════════════════════════════════

def style_cell(cell, bold=False, bg=None, fg="000000",
               size=10, align="left", border=False):
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border:
        s = Side(style="thin", color="D0E8D8")
        cell.border = Border(left=s, right=s, top=s, bottom=s)


def write_title(ws, row, titre, nb_cols=2):
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=titre)
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=12, align="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    return row + 1


def write_headers(ws, row, cols):
    ws.row_dimensions[row].height = 22
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=col)
        style_cell(cell, bold=True, bg=C_VERT_MOYEN, fg=C_BLANC,
                   size=10, align="center", border=True)
    return row + 1


def write_data(ws, row, df):
    for i, r in enumerate(df.itertuples(index=False)):
        ws.row_dimensions[row].height = 18
        bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
        for c, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c, value=val)
            style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
        row += 1
    return row


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def onglet_resume(wb, df):
    ws = wb.active
    ws.title = "Résumé général"
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=1,
                value="RAPPORT QAQC - TABLE MÉNAGE - RGPH-5 Sénégal")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=15, align="center")
    ws.merge_cells("A1:B1")
    row += 2

    row = write_title(ws, row, "STATISTIQUES GÉNÉRALES")
    row = write_headers(ws, row, ["Indicateur", "Valeur"])
    row = write_data(ws, row, stats_generales(df))
    row += 1

    if "taille_menage" in df.columns:
        row = write_title(ws, row, "TAILLE DU MÉNAGE")
        row = write_headers(ws, row, ["Indicateur", "Valeur"])
        row = write_data(ws, row, stats_taille(df))
        row += 1

    if "age_cm" in df.columns:
        row = write_title(ws, row, "ÂGE DU CHEF DE MÉNAGE")
        row = write_headers(ws, row, ["Indicateur", "Valeur"])
        row = write_data(ws, row, stats_age_cm(df))

    set_widths(ws, [45, 35])


def onglet_dist(wb, df, col, nom_onglet, titre):
    if col not in df.columns:
        return
    ws = wb.create_sheet(nom_onglet)
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=titre)
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=13, align="center")
    ws.merge_cells(f"A{row}:C{row}")
    row += 2

    dist = dist_variable(df, col)
    row = write_headers(ws, row, ["Modalité", "Effectif", "Proportion (%)"])
    row = write_data(ws, row, dist)

    ws.row_dimensions[row].height = 20
    for c_i, val in enumerate(["TOTAL", f"{len(df):,}", "100.00"], 1):
        cell = ws.cell(row=row, column=c_i, value=val)
        style_cell(cell, bold=True, bg=C_VERT_FONCE, fg=C_BLANC,
                   size=10, border=True)
    set_widths(ws, [40, 15, 18])


def onglet_qualite(wb, df, df_log):
    ws = wb.create_sheet("Qualité des données")
    ws.sheet_view.showGridLines = False

    row = 1
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1,
                value="QUALITÉ DES DONNÉES - VALEURS MANQUANTES")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=13, align="center")
    ws.merge_cells(f"A{row}:D{row}")
    row += 2

    manq = stats_manquants(df)
    row = write_headers(ws, row, ["Variable", "Nb manquants", "Taux (%)", "Statut"])

    for i, r in enumerate(manq.itertuples(index=False)):
        ws.row_dimensions[row].height = 18
        bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
        for c_i, val in enumerate(r, 1):
            cell = ws.cell(row=row, column=c_i, value=val)
            if c_i == 4:
                if val == "Complet":
                    style_cell(cell, bold=True, bg="C6F6D5", fg="1A4731",
                               size=10, align="center", border=True)
                elif val == "Faible < 5%":
                    style_cell(cell, bg="FEFCBF", fg="744210",
                               size=10, align="center", border=True)
                else:
                    style_cell(cell, bg="FED7D7", fg="742A2A",
                               size=10, align="center", border=True)
            else:
                style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
        row += 1

    if df_log is not None and len(df_log) > 0:
        row += 2
        ws.row_dimensions[row].height = 28
        c = ws.cell(row=row, column=1, value="LOG DES CORRECTIONS EFFECTUÉES")
        style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=12, align="center")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        row = write_headers(ws, row, ["Contrôle", "Nb cas", "Action", ""])
        df_log_disp = df_log[["controle", "nb_cas", "action"]].copy()
        df_log_disp[""] = ""
        row = write_data(ws, row, df_log_disp)

    set_widths(ws, [42, 12, 55, 5])


def onglet_notes(wb):
    ws = wb.create_sheet("Notes méthodologiques")
    ws.sheet_view.showGridLines = False

    notes = [
        ("TRAITEMENT DES VALEURS ABERRANTES", [
            ("Âge du CM (age_cm)",
             "Bornes définies : [0, 120] ans. Toute valeur hors de ces bornes est remplacée par NaN. Justification : aucun être humain ne peut avoir un âge négatif ou supérieur à 120 ans."),
            ("Taille du ménage (taille_menage)",
             "Seuil fixé au 99e percentile observé = 125 membres. Valeurs > 125 remplacées par NaN. Justification empirique : distribution observée - P50=12, P75=46, P90=78, P99=125."),
        ]),
        ("CONTRÔLES DE COHÉRENCE", [
            ("Sexe vs situation matrimoniale",
             "Un CM de sexe masculin ne peut pas être déclaré épouse en polygamie. Action : situation_matrimoniale_cm - NaN."),
        ]),
    ]

    row = 1
    ws.row_dimensions[row].height = 40
    c = ws.cell(row=row, column=1,
                value="NOTES MÉTHODOLOGIQUES - DÉCISIONS DE TRAITEMENT")
    style_cell(c, bold=True, bg=C_VERT_FONCE, fg=C_BLANC, size=14, align="center")
    ws.merge_cells(f"A{row}:B{row}")
    row += 2

    for section_titre, items in notes:
        ws.row_dimensions[row].height = 26
        c = ws.cell(row=row, column=1, value=section_titre)
        style_cell(c, bold=True, bg=C_VERT_MOYEN, fg=C_BLANC,
                   size=11, align="left")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        row = write_headers(ws, row, ["Point", "Explication"])

        for i, (point, explication) in enumerate(items):
            ws.row_dimensions[row].height = 45
            bg = C_VERT_CLAIR if i % 2 == 0 else C_BLANC
            for c_i, val in enumerate([point, explication], 1):
                cell = ws.cell(row=row, column=c_i, value=val)
                style_cell(cell, bg=bg, fg=C_TEXTE, size=10, border=True)
            row += 1
        row += 1

    set_widths(ws, [35, 80])


def generer_rapport_excel(df, df_log, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    onglet_resume(wb, df)
    onglet_dist(wb, df, "sexe_cm", "Sexe CM", "RÉPARTITION PAR SEXE DU CM")
    onglet_dist(wb, df, "niveau_etudes_cm", "Niveau études CM", "NIVEAU D'ÉTUDES DU CM")
    onglet_dist(wb, df, "situation_matrimoniale_cm", "Situation matrimoniale", "SITUATION MATRIMONIALE DU CM")
    onglet_dist(wb, df, "scolarisation_cm", "Scolarisation CM", "SCOLARISATION DU CM")
    onglet_dist(wb, df, "statut_emploi_cm", "Statut emploi CM", "STATUT D'EMPLOI DU CM")
    onglet_dist(wb, df, "secteur_instit_cm", "Secteur activité CM", "SECTEUR D'ACTIVITÉ DU CM (ISIC Rev 4)")
    onglet_dist(wb, df, "region", "Régions", "DISTRIBUTION DES MÉNAGES PAR RÉGION")
    
    milieu_col = "milieu" if "milieu" in df.columns else ("milieu_residence" if "milieu_residence" in df.columns else None)
    if milieu_col:
        onglet_dist(wb, df, milieu_col, "Milieu", "DISTRIBUTION DES MÉNAGES PAR MILIEU")
        
    onglet_qualite(wb, df, df_log)
    onglet_notes(wb)

    wb.save(output_path)
    print(f"  Rapport Excel : {output_path}")


# ════════════════════════════════════════════════════════════════
# RAPPORT HTML
# ════════════════════════════════════════════════════════════════

def df_to_html(df):
    if len(df) == 0:
        return "<p style='color:#718096;'>Aucune donnée disponible</p>"
    rows_html = ""
    for i, row in enumerate(df.itertuples(index=False)):
        bg = "#f0fff4" if i % 2 == 0 else "#ffffff"
        cells = "".join(
            f'<td style="padding:10px 14px;border-bottom:1px solid #c6f6d5;'
            f'color:#1c3829;">{val}</td>'
            for val in row
        )
        rows_html += f'<tr style="background:{bg};">{cells}</tr>'

    headers = "".join(
        f'<th style="padding:12px 14px;text-align:left;font-weight:600;'
        f'color:#ffffff;background:#276749;letter-spacing:0.03em;">{col}</th>'
        for col in df.columns
    )
    return f"""
    <table style="width:100%;border-collapse:collapse;border-radius:8px;
                  overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,0.07);">
         <thead><tr>{headers}</tr></thead>
         <tbody>{rows_html}</tbody>
    </table>"""


def carte(label, valeur, couleur="#1A4731"):
    return f"""
    <div style="background:#ffffff;border-radius:10px;padding:20px 22px;
                box-shadow:0 2px 6px rgba(0,0,0,0.06);
                border-left:4px solid {couleur};flex:1;min-width:160px;">
        <div style="font-size:10px;text-transform:uppercase;letter-spacing:0.08em;
                    color:#4a5568;font-weight:700;margin-bottom:6px;">{label}</div>
        <div style="font-size:22px;font-weight:800;color:{couleur};">{valeur}</div>
    </div>"""


def section(titre, contenu):
    return f"""
    <div style="margin-bottom:32px;">
         <h2 style="font-size:13px;font-weight:700;color:#1A4731;
                    text-transform:uppercase;letter-spacing:0.07em;
                    border-bottom:2px solid #276749;padding-bottom:7px;
                    margin-bottom:14px;">{titre}</h2>
         {contenu}
    </div>"""


def note_item(point, explication):
    return f"""
    <div style="display:grid;grid-template-columns:220px 1fr;gap:0;
                border-bottom:1px solid #c6f6d5;">
        <div style="padding:12px 14px;background:#f0fff4;
                    font-weight:600;color:#1A4731;font-size:13px;
                    border-right:2px solid #276749;">{point}</div>
        <div style="padding:12px 16px;color:#2d3748;
                    font-size:13px;line-height:1.6;">{explication}</div>
    </div>"""


def bloc_note(titre_bloc, items):
    items_html = "".join(note_item(p, e) for p, e in items)
    return f"""
    <div style="margin-bottom:24px;border-radius:8px;overflow:hidden;
                box-shadow:0 1px 4px rgba(0,0,0,0.07);">
        <div style="background:#276749;color:#ffffff;padding:10px 16px;
                    font-size:12px;font-weight:700;letter-spacing:0.05em;
                    text-transform:uppercase;">{titre_bloc}</div>
        {items_html}
    </div>"""


def generer_rapport_html(df, df_log, output_path):
    milieu_col = "milieu" if "milieu" in df.columns else ("milieu_residence" if "milieu_residence" in df.columns else None)
    
    milieu_urb_pct = f"{(df[milieu_col]=='Urbain').mean()*100:.1f}%" if milieu_col else "N/A"
    milieu_rur_pct = f"{(df[milieu_col]=='Rural').mean()*100:.1f}%" if milieu_col else "N/A"
    
    sexe_m_pct = f"{(df['sexe_cm']=='Masculin').mean()*100:.1f}%" if "sexe_cm" in df.columns else "N/A"
    sexe_f_pct = f"{(df['sexe_cm']=='Féminin').mean()*100:.1f}%" if "sexe_cm" in df.columns else "N/A"
    
    age_med = f"{df['age_cm'].median():.0f} ans" if "age_cm" in df.columns else "N/A"
    taille_med = f"{df['taille_menage'].median():.0f}" if "taille_menage" in df.columns else "N/A"
    
    # Cartes
    cartes = f"""
    <div style="display:flex;gap:14px;flex-wrap:wrap;margin-bottom:8px;">
        {carte("Ménages traités",       f"{len(df):,}",                                       "#1A4731")}
        {carte("Milieu urbain",         milieu_urb_pct,                                       "#276749")}
        {carte("Milieu rural",          milieu_rur_pct,                                       "#276749")}
        {carte("CM masculin",           sexe_m_pct,                                           "#2F855A")}
        {carte("CM féminin",            sexe_f_pct,                                           "#2F855A")}
        {carte("Âge médian CM",         age_med,                                              "#38A169")}
        {carte("Taille médian ménage", taille_med,                                            "#48BB78")}
    </div>"""

    # Notes méthodologiques
    notes_html = "".join([
        bloc_note("Traitement des valeurs aberrantes", [
            ("Âge du CM",
             "Bornes définies : [0, 120] ans. Toute valeur hors de ces bornes est remplacée par NaN."),
            ("Taille du ménage",
             "Seuil fixé au 99e percentile observé = 125 membres. Valeurs > 125 remplacées par NaN. Distribution observée : P50=12, P75=46, P90=78, P99=125."),
        ]),
        bloc_note("Contrôles de cohérence", [
            ("Sexe vs situation matrimoniale",
             "Un CM masculin ne peut pas être déclaré épouse en polygamie. Situation matrimoniale mise à NaN en cas d'incohérence."),
        ]),
    ])

    # Grille de tables dynamiques
    sections_dynamiques = []
    
    # Taille et Age
    t_html = ""
    if "taille_menage" in df.columns:
        t_html += f'<div class="card">{section("Taille du ménage", df_to_html(stats_taille(df)))}</div>'
    if "age_cm" in df.columns:
        t_html += f'<div class="card">{section("Âge du chef de ménage", df_to_html(stats_age_cm(df)))}</div>'
    if t_html:
        sections_dynamiques.append(f'<div class="grid-2">{t_html}</div><br>')
        
    # Sexe et Milieu
    sm_html = ""
    if "sexe_cm" in df.columns:
        sm_html += f'<div class="card">{section("Sexe du CM", df_to_html(dist_variable(df, "sexe_cm")))}</div>'
    if milieu_col:
        sm_html += f'<div class="card">{section("Milieu de résidence", df_to_html(dist_variable(df, milieu_col)))}</div>'
    if sm_html:
        sections_dynamiques.append(f'<div class="grid-2">{sm_html}</div><br>')
        
    # Scolarisation et Situation matrimoniale
    ss_html = ""
    if "scolarisation_cm" in df.columns:
        ss_html += f'<div class="card">{section("Scolarisation du CM", df_to_html(dist_variable(df, "scolarisation_cm")))}</div>'
    if "situation_matrimoniale_cm" in df.columns:
        ss_html += f'<div class="card">{section("Situation matrimoniale du CM", df_to_html(dist_variable(df, "situation_matrimoniale_cm")))}</div>'
    if ss_html:
        sections_dynamiques.append(f'<div class="grid-2">{ss_html}</div><br>')

    # Niveau études
    if "niveau_etudes_cm" in df.columns:
        lbl_etudes = "Niveau d'études du CM"
        sections_dynamiques.append(f'<div class="card">{section(lbl_etudes, df_to_html(dist_variable(df, "niveau_etudes_cm")))}</div><br>')
        
    # Statut Emploi et Secteur d'Activité
    emp_html = ""
    if "statut_emploi_cm" in df.columns:
        lbl_emploi = "Statut d'emploi du CM"
        emp_html += f'<div class="card">{section(lbl_emploi, df_to_html(dist_variable(df, "statut_emploi_cm")))}</div>'
    if "secteur_instit_cm" in df.columns:
        lbl_secteur = "Secteur d'activité du CM (ISIC)"
        emp_html += f'<div class="card">{section(lbl_secteur, df_to_html(dist_variable(df, "secteur_instit_cm")))}</div>'
    if emp_html:
        sections_dynamiques.append(f'<div class="grid-2">{emp_html}</div><br>')
        
    # Région
    if "region" in df.columns:
        lbl_region = "Distribution par région"
        sections_dynamiques.append(f'<div class="card">{section(lbl_region, df_to_html(dist_variable(df, "region")))}</div><br>')

    dinamic_content = "\n".join(sections_dynamiques)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>QAQC - Table Ménage - RGPH-5</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
      font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
      background: #f7faf7;
      color: #2d3748;
      line-height: 1.6;
  }}
  .header {{
      background: linear-gradient(135deg, #1A4731 0%, #276749 60%, #38A169 100%);
      color: white;
      padding: 44px 52px;
  }}
  .header h1 {{
      font-size: 26px;
      font-weight: 800;
      letter-spacing: -0.02em;
      margin-bottom: 8px;
  }}
  .header p {{ font-size: 14px; opacity: 0.85; margin-top: 4px; }}
  .badge {{
      display: inline-block;
      background: rgba(255,255,255,0.18);
      border: 1px solid rgba(255,255,255,0.35);
      border-radius: 20px;
      padding: 5px 16px;
      font-size: 12px;
      font-weight: 700;
      margin-top: 14px;
      letter-spacing: 0.05em;
  }}
  .container {{
      max-width: 1120px;
      margin: 0 auto;
      padding: 44px 36px;
  }}
  .grid-2 {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
  }}
  .card {{
      background: white;
      border-radius: 12px;
      padding: 26px;
      box-shadow: 0 1px 4px rgba(0,0,0,0.06);
  }}
  .footer {{
      text-align: center;
      padding: 22px;
      font-size: 12px;
      color: #718096;
      border-top: 1px solid #c6f6d5;
      margin-top: 40px;
  }}
</style>
</head>
<body>

<div class="header">
    <h1>Rapport QAQC - Table Ménage</h1>
    <p>5ème Recensement Général de la Population et de l'Habitat - Sénégal</p>
    <p>Échantillon 1/10 de la base nationale</p>
    <span class="badge">Groupe 2 - Pipeline de traitement</span>
</div>

<div class="container">

    {section("Vue d'ensemble", cartes)}

    {dinamic_content}

    <div class="card">
        {section("Qualité des données - Valeurs manquantes", df_to_html(stats_manquants(df)))}
    </div><br>

    <div class="card">
        {section("Notes méthodologiques - Décisions de traitement", notes_html)}
    </div>

</div>

<div class="footer">
    Généré automatiquement par le pipeline RGPH-5 - Groupe 2
</div>

</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")
    print(f"  Rapport HTML  : {output_path}")


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════

def main():
    section_header("ÉTAPE 5 - RAPPORT QAQC TABLE MÉNAGE")

    hh_path = os.path.join(OUTPUT_DIR, "hh_final.csv")
    if not os.path.exists(hh_path):
        hh_path = os.path.join(OUTPUT_DIR, "rgph5_hh_clean.csv")
        
    if not os.path.exists(hh_path):
        raise FileNotFoundError(
            f"Introuvable : {hh_path}\n- Lancer d'abord le nettoyage ou la fusion"
        )
        
    df = pd.read_csv(hh_path)
    print(f"\n  Table ménage chargée : {len(df):,} lignes × {df.shape[1]} colonnes")

    df_log = pd.DataFrame(columns=["controle", "nb_cas", "action"])

    qaqc_dir = QAQC_DIR
    os.makedirs(qaqc_dir, exist_ok=True)

    print("\n[Génération des rapports ménages]")
    generer_rapport_excel(df, df_log, Path(os.path.join(qaqc_dir, "qaqc_menage.xlsx")))
    generer_rapport_html(df, df_log,  Path(os.path.join(qaqc_dir, "qaqc_menage.html")))

    print(f"\n  Rapports disponibles dans : {qaqc_dir}")


if __name__ == "__main__":
    main()
