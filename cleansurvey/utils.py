"""
utils.py
=========
Fonctions utilitaires réutilisables pour le pipeline RGPH.

Toutes les fonctions sont conçues pour être indépendantes du recensement :
elles ne font référence à aucun nom de variable brute. Seul config.py
contient les références aux variables spécifiques au recensement traité.
"""

import sys
import json
import pathlib
import warnings
from typing import Dict, List, Optional, Tuple, Any

import numpy as np
import pandas as pd
import pyreadstat
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)


# ══════════════════════════════════════════════════════════════════════════════
# 1. CHARGEMENT DES DONNÉES
# ══════════════════════════════════════════════════════════════════════════════

def load_sav(path: pathlib.Path,
             row_limit: Optional[int] = None,
             apply_formats: bool = False,
             cols: Optional[List[str]] = None) -> Tuple[pd.DataFrame, Any]:
    """
    Charge un fichier SPSS .sav avec pyreadstat.

    Parameters
    ----------
    path         : chemin vers le fichier .sav
    row_limit    : nombre max de lignes à lire (None = tout)
    apply_formats: appliquer les étiquettes de valeur SPSS (True) ou
                   conserver les codes numériques (False)
    cols         : liste de colonnes à sélectionner (None = toutes)

    Returns
    -------
    (df, meta)   : DataFrame et objet métadonnées pyreadstat
    """
    kwargs = dict(
        apply_value_formats=apply_formats,
    )
    if row_limit:
        kwargs["row_limit"] = row_limit
    if cols:
        kwargs["usecols"] = cols

    print(f"[load_sav] Chargement de {path.name} ...")
    df, meta = pyreadstat.read_sav(str(path), **kwargs)
    print(f"  >> {df.shape[0]:,} lignes x {df.shape[1]} colonnes chargees")
    return df, meta


def get_metadata(path: pathlib.Path) -> Any:
    """Lit uniquement les métadonnées (très rapide, sans charger les données)."""
    _, meta = pyreadstat.read_sav(str(path), metadataonly=True)
    return meta


# ══════════════════════════════════════════════════════════════════════════════
# 2. RENOMMAGE ET SÉLECTION DES VARIABLES
# ══════════════════════════════════════════════════════════════════════════════

def select_and_rename(df: pd.DataFrame,
                      rename_map: Dict[str, str],
                      keep_unmapped: bool = False) -> pd.DataFrame:
    """
    Sélectionne et renomme les colonnes selon rename_map.

    Parameters
    ----------
    df           : DataFrame source
    rename_map   : {nom_brut: nom_normalise}
    keep_unmapped: garder les colonnes absentes de rename_map

    Returns
    -------
    DataFrame avec uniquement les colonnes sélectionnées et renommées
    """
    # Colonnes présentes dans le DataFrame et dans le mapping
    present = {k: v for k, v in rename_map.items() if k in df.columns}
    absent  = [k for k in rename_map if k not in df.columns]

    if absent:
        print(f"  [WARN] Variables absentes du fichier : {absent}")

    if keep_unmapped:
        # Garder tout + renommer ce qui est dans le mapping
        return df.rename(columns=present)
    else:
        # Sélectionner uniquement les colonnes dans le mapping
        df_out = df[list(present.keys())].rename(columns=present)
        return df_out


# ══════════════════════════════════════════════════════════════════════════════
# 3. TRAITEMENT DES VALEURS MANQUANTES
# ══════════════════════════════════════════════════════════════════════════════

def flag_and_nullify_missing(df: pd.DataFrame,
                             missing_codes: Dict[str, List],
                             flag_col_suffix: str = "_flag_miss") -> pd.DataFrame:
    """
    Pour chaque variable, remplace les codes "manquant" par NaN et
    crée une colonne indicatrice de valeur manquante originellement.

    Parameters
    ----------
    df              : DataFrame
    missing_codes   : {nom_colonne: [codes_manquants]}
    flag_col_suffix : suffixe pour les colonnes indicatrices

    Returns
    -------
    DataFrame enrichi avec les colonnes flag et les codes remplacés par NaN
    """
    df = df.copy()
    for col, codes in missing_codes.items():
        if col not in df.columns:
            continue
        mask = df[col].isin(codes)
        flag_col = col + flag_col_suffix
        df[flag_col] = mask.astype(int)
        df.loc[mask, col] = np.nan
    return df


def missing_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Retourne un tableau résumé des valeurs manquantes par colonne.

    Returns
    -------
    DataFrame avec colonnes : variable, n_total, n_manquants, pct_manquants
    """
    total = len(df)
    rows  = []
    for col in df.columns:
        if col.endswith("_flag_miss"):
            continue
        n_miss = df[col].isna().sum()
        rows.append({
            "variable"      : col,
            "n_total"       : total,
            "n_manquants"   : int(n_miss),
            "pct_manquants" : round(100 * n_miss / total, 2) if total > 0 else 0,
        })
    return pd.DataFrame(rows).sort_values("pct_manquants", ascending=False)


# ══════════════════════════════════════════════════════════════════════════════
# 4. TRAITEMENT DES VALEURS ABERRANTES
# ══════════════════════════════════════════════════════════════════════════════

def flag_outliers(df: pd.DataFrame,
                  valid_ranges: Dict[str, Tuple[float, float]],
                  action: str = "nullify",
                  flag_col_suffix: str = "_flag_outlier") -> pd.DataFrame:
    """
    Détecte et traite les valeurs hors plage pour les variables numériques.

    Parameters
    ----------
    df              : DataFrame
    valid_ranges    : {nom_col: (min_val, max_val)}
    action          : "nullify" (remplacer par NaN) ou "flag_only"
    flag_col_suffix : suffixe pour les colonnes indicatrices

    Returns
    -------
    DataFrame enrichi
    """
    df = df.copy()
    for col, (vmin, vmax) in valid_ranges.items():
        if col not in df.columns:
            continue
        mask = (df[col] < vmin) | (df[col] > vmax)
        flag_col = col + flag_col_suffix
        df[flag_col] = mask.astype(int)
        if action == "nullify":
            df.loc[mask, col] = np.nan
    return df


def outlier_summary(df: pd.DataFrame,
                    flag_col_suffix: str = "_flag_outlier") -> pd.DataFrame:
    """
    Résume les valeurs aberrantes détectées.

    Returns
    -------
    DataFrame : variable | n_aberrants | pct_aberrants
    """
    rows = []
    total = len(df)
    flag_cols = [c for c in df.columns if c.endswith(flag_col_suffix)]
    for fc in flag_cols:
        var = fc.replace(flag_col_suffix, "")
        n   = int(df[fc].sum())
        rows.append({
            "variable"     : var,
            "n_aberrants"  : n,
            "pct_aberrants": round(100 * n / total, 2) if total > 0 else 0,
        })
    return pd.DataFrame(rows).sort_values("n_aberrants", ascending=False)


# ══════════════════════════════════════════════════════════════════════════════
# 5. RECODAGE DES MODALITÉS
# ══════════════════════════════════════════════════════════════════════════════

def apply_recodings(df: pd.DataFrame,
                    recodings: Dict[str, Dict]) -> pd.DataFrame:
    """
    Applique les mappings de recodage (code numérique -> étiquette texte).

    Parameters
    ----------
    df        : DataFrame
    recodings : {nom_col: {code: etiquette}}

    Returns
    -------
    DataFrame recodé
    """
    df = df.copy()
    for col, mapping in recodings.items():
        if col not in df.columns:
            continue
        df[col] = df[col].map(mapping).where(df[col].notna(), np.nan)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 6. DÉRIVATION DE VARIABLES
# ══════════════════════════════════════════════════════════════════════════════

def compute_groupe_age(df: pd.DataFrame,
                       age_col: str,
                       bins: List[int],
                       labels: List[str],
                       out_col: str = "groupe_age") -> pd.DataFrame:
    """Crée une variable groupes d'âge à partir de l'âge révolu."""
    df = df.copy()
    if age_col not in df.columns:
        print(f"  [WARN] Colonne age '{age_col}' absente.")
        return df
    df[out_col] = pd.cut(df[age_col], bins=bins, labels=labels, right=False)
    return df


def compute_alphabetisation(df: pd.DataFrame,
                             alpha_cols: List[str],
                             alpha_oui_code,
                             out_col: str = "alphabetise") -> pd.DataFrame:
    """
    Crée une variable binaire 'alphabetise' = True si alphabétisé
    dans au moins une langue.

    Parameters
    ----------
    alpha_cols     : liste des colonnes d'alphabétisation (une par langue)
    alpha_oui_code : valeur représentant "Oui" dans ces colonnes
    """
    df = df.copy()
    present = [c for c in alpha_cols if c in df.columns]
    if not present:
        print(f"  [WARN] Aucune colonne d'alphabetisation trouvee.")
        return df
    # True si au moins une langue = code "Oui"
    df[out_col] = df[present].apply(
        lambda row: any(v == alpha_oui_code for v in row if pd.notna(v)),
        axis=1
    ).astype(int)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# 7. CONTRÔLES DE COHÉRENCE
# ══════════════════════════════════════════════════════════════════════════════

def run_coherence_checks(df: pd.DataFrame,
                         checks: List[Dict]) -> pd.DataFrame:
    """
    Exécute les contrôles de cohérence et retourne un tableau de résultats.

    Parameters
    ----------
    checks : liste de dicts {nom, description, condition (expression pandas)}

    Returns
    -------
    DataFrame : nom | description | n_violations | pct_violations
    """
    results = []
    total   = len(df)
    for chk in checks:
        try:
            mask     = df.eval(chk["condition"])
            n_viol   = int(mask.sum())
            pct_viol = round(100 * n_viol / total, 2) if total > 0 else 0
        except Exception as e:
            n_viol   = -1
            pct_viol = -1
            print(f"  [WARN] Controle '{chk['nom']}' non evalue : {e}")
        results.append({
            "nom"           : chk["nom"],
            "description"   : chk["description"],
            "n_violations"  : n_viol,
            "pct_violations": pct_viol,
        })
    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
# 8. DISTRIBUTIONS (pour le QAQC)
# ══════════════════════════════════════════════════════════════════════════════

def compute_distribution(df: pd.DataFrame,
                         col: str,
                         normalize: bool = True) -> pd.DataFrame:
    """
    Calcule la distribution (effectifs et %) d'une variable catégorielle.
    """
    if col not in df.columns:
        return pd.DataFrame()
    vc = df[col].value_counts(dropna=False)
    result = pd.DataFrame({"modalite": vc.index, "effectif": vc.values})
    if normalize:
        result["pourcentage"] = (result["effectif"] / result["effectif"].sum() * 100).round(2)
    result.insert(0, "variable", col)
    return result


def compute_all_distributions(df: pd.DataFrame,
                               cat_cols: List[str]) -> pd.DataFrame:
    """Calcule les distributions de toutes les colonnes catégorielles listées."""
    frames = [compute_distribution(df, col) for col in cat_cols if col in df.columns]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# 9. GÉNÉRATION DU FICHIER QAQC EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _style_header(ws, row: int = 1,
                  fill_color: str = "1F4E79",
                  font_color: str = "FFFFFF"):
    """Stylise la ligne d'en-tête d'une feuille Excel."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(color=font_color, bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def _auto_col_width(ws, min_width: int = 12, max_width: int = 45):
    """Ajuste automatiquement la largeur des colonnes."""
    for col_cells in ws.columns:
        length = max(
            len(str(c.value)) if c.value is not None else 0
            for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(length + 2, max_width))


def _df_to_sheet(wb: openpyxl.Workbook,
                 sheet_name: str,
                 df: pd.DataFrame,
                 fill_color: str = "1F4E79"):
    """Écrit un DataFrame dans un onglet Excel avec mise en forme."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    # En-têtes
    ws.append(list(df.columns))
    _style_header(ws, row=1, fill_color=fill_color)
    # Données
    for row in df.itertuples(index=False):
        ws.append(list(row))
    # Zebra stripes
    fill_light = PatternFill(start_color="EFF3FB", end_color="EFF3FB", fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if i % 2 == 0:
            for cell in row:
                cell.fill = fill_light
    _auto_col_width(ws)


def generate_qaqc(
    df_clean         : pd.DataFrame,
    df_missing       : pd.DataFrame,
    df_outliers      : pd.DataFrame,
    df_distributions : pd.DataFrame,
    df_coherence     : pd.DataFrame,
    df_estimations   : pd.DataFrame,
    output_path      : pathlib.Path,
    n_raw            : int = 0,
):
    """
    Génère le fichier QAQC Excel avec 6 onglets.

    Parameters
    ----------
    df_clean        : DataFrame nettoyé final
    df_missing      : résumé des valeurs manquantes
    df_outliers     : résumé des valeurs aberrantes
    df_distributions: distributions des variables catégorielles
    df_coherence    : résultats des contrôles de cohérence
    df_estimations  : indicateurs primaires
    output_path     : chemin de sortie .xlsx
    n_raw           : nombre de lignes brutes avant nettoyage
    """
    import datetime
    wb = openpyxl.Workbook()

    # ── Onglet 1 : Résumé ────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resume"
    infos = [
        ("Date de traitement"         , datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Fichier source"             , "dixieme_RGPH_5_indiv_SECTION_B.sav"),
        ("N lignes brutes"            , n_raw),
        ("N lignes apres nettoyage"   , len(df_clean)),
        ("N variables produites"      , df_clean.shape[1]),
        ("N variables avec manquants" , int((df_missing["n_manquants"] > 0).sum())),
        ("N controles coherence"      , len(df_coherence)),
        ("N violations coherence"     , int(df_coherence["n_violations"].clip(lower=0).sum())),
    ]
    ws_sum.append(["Indicateur", "Valeur"])
    _style_header(ws_sum, row=1, fill_color="1F4E79")
    for row in infos:
        ws_sum.append(list(row))
    _auto_col_width(ws_sum)

    # ── Onglets 2-6 ──────────────────────────────────────────────────────────
    _df_to_sheet(wb, "Manquants"          , df_missing      , fill_color="C55A11")
    _df_to_sheet(wb, "Aberrants"          , df_outliers      , fill_color="843C0C")
    _df_to_sheet(wb, "Distributions"      , df_distributions , fill_color="1F4E79")
    _df_to_sheet(wb, "Coherence"          , df_coherence     , fill_color="375623")
    _df_to_sheet(wb, "Estimations"        , df_estimations   , fill_color="7030A0")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  >> QAQC Excel sauvegarde : {output_path}")
