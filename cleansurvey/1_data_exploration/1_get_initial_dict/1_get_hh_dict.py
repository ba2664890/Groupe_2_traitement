"""
1_data_exploration/1_get_initial_dict/1_get_hh_dict.py
=======================================================
Étape 1 - Dictionnaire initial des variables ménage.

Génère un fichier Excel listant toutes les variables disponibles
dans habitat_SECTION_E.sav et indiv_SECTION_B.sav (pour le CM),
avec leur type, nombre de valeurs uniques et taux de manquants.

Output : output/dict/hh_dict.xlsx
"""
# C'est la PREMIÈRE étape du pipeline - l'étape d'exploration.
# Avant de sélectionner ou nettoyer quoi que ce soit, on commence
# par dresser un inventaire complet de toutes les variables disponibles.
# Le résultat est un fichier Excel : hh_dict.xlsx
# Ce fichier sert de référence pour comprendre la base et
# justifier les choix de variables dans le rapport.


# -- IMPORTS ------------------------------------------------------

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
# Ce fichier est dans un sous-dossier profond :
# cleansurvey/1_data_exploration/1_get_initial_dict/
# .parents[2] remonte de 2 niveaux pour atteindre cleansurvey/
# et trouver config.py et utils.py

import pandas as pd
import numpy as np
import openpyxl
# openpyxl est la bibliothèque pour créer et modifier des fichiers Excel (.xlsx)
# Elle permet de mettre en forme les cellules (couleurs, polices, largeurs...)

from openpyxl.styles import Font, PatternFill, Alignment
# Font       : style de police (gras, couleur du texte...)
# PatternFill: couleur de fond des cellules
# Alignment  : alignement du texte (centré, gauche...)

from config import FILE_HABITAT, FILE_INDIV, OUTPUT_DIR
# Chemins vers les fichiers sources et le dossier de sortie

from utils import read_sav, section_header
# read_sav       : lit les .sav avec gestion mémoire (chunks)
# section_header : affiche les titres dans le terminal


# ════════════════════════════════════════════════════════════════
# FONCTION 1 - build_dict()
# Construire le dictionnaire des variables d'une table
# ════════════════════════════════════════════════════════════════

def build_dict(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """
    Construit le dictionnaire d'une table :
    variable, source, type, nb_valeurs_uniques, taux_manquant, min, max, exemple.
    """
    # Un dictionnaire de variables = un tableau qui décrit chaque colonne.
    # Pour chaque colonne de la base, on calcule des statistiques descriptives.
    # C'est l'équivalent de proc contents en SAS ou str() en R.

    rows = []
    # Liste vide qui va accumuler une ligne de description par variable

    for col in df.columns:
        # On parcourt toutes les colonnes du DataFrame une par une

        serie = df[col]
        # serie = la colonne entière (un objet pandas Series)

        rows.append({
            "variable": col,
            # Nom de la colonne (ex : "A06", "B33", "AGE_CM"...)

            "source": source,
            # Nom du fichier d'origine (ex : "habitat_SECTION_E.sav")
            # Utile pour savoir d'où vient chaque variable

            "type": str(serie.dtype),
            # Type de données pandas : float64, object, int64...
            # float64 = nombre décimal, object = texte, int64 = entier

            "nb_valeurs_uniques": serie.nunique(dropna=True),
            # Compte le nombre de valeurs distinctes dans la colonne.
            # dropna=True : on exclut les NaN du comptage
            # Ex : A10 (milieu) - 2 valeurs uniques (1 et 2)
            # Ex : A06 (district) - 21 235 valeurs uniques

            "taux_manquant_%": round(serie.isna().mean() * 100, 2),
            # serie.isna()  - série de True/False (True = NaN)
            # .mean()        - proportion de NaN (entre 0 et 1)
            # * 100          - conversion en pourcentage
            # round(..., 2)  - arrondi à 2 décimales
            # Ex : 60.95% de manquants sur niveau_etude_cm

            "min": serie.min() if pd.api.types.is_numeric_dtype(serie) else "",
            # Valeur minimale - calculée UNIQUEMENT si la colonne est numérique
            # pd.api.types.is_numeric_dtype() vérifie le type
            # Pour les colonnes texte (object) - on laisse vide ""

            "max": serie.max() if pd.api.types.is_numeric_dtype(serie) else "",
            # Valeur maximale - même logique que min

            "exemple_valeur": serie.dropna().iloc[0] if serie.dropna().shape[0] > 0 else "",
            # Prend la première valeur non-NaN de la colonne comme exemple.
            # serie.dropna()     - supprime les NaN
            # .iloc[0]           - prend le premier élément
            # serie.dropna().shape[0] > 0 : vérifie qu'il reste au moins une valeur
            # Si toute la colonne est NaN - exemple vide ""
        })

    return pd.DataFrame(rows)
    # Convertit la liste de dictionnaires en DataFrame :
    # chaque dictionnaire = une ligne = une variable de la base
    # Résultat : un tableau avec 8 colonnes de description


# ════════════════════════════════════════════════════════════════
# FONCTION 2 - save_dict_excel()
# Sauvegarder le dictionnaire en fichier Excel mis en forme
# ════════════════════════════════════════════════════════════════

def save_dict_excel(df_dict: pd.DataFrame, output_path: Path) -> None:
    """Sauvegarde le dictionnaire dans un fichier Excel mis en forme."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Crée le dossier output/dict/ s'il n'existe pas encore

    wb = openpyxl.Workbook()
    # Crée un nouveau classeur Excel vide

    ws = wb.active
    # Sélectionne la feuille active (la première feuille par défaut)

    ws.title = "Dictionnaire ménage"
    # Renomme l'onglet

    fill_h = PatternFill("solid", fgColor="2D6A9F")
    # Définit un remplissage bleu foncé (#2D6A9F) pour les en-têtes
    # "solid" = remplissage plein (pas de motif)

    font_h = Font(bold=True, color="FFFFFF")
    # Police en gras blanc pour les en-têtes
    # Combiné avec le fond bleu - contraste lisible

    # -- Écriture des en-têtes (ligne 1) --
    for c, col in enumerate(df_dict.columns, 1):
        # enumerate(..., 1) : parcourt les colonnes en commençant à 1
        # (Excel commence à 1, pas à 0 comme Python)

        cell = ws.cell(row=1, column=c, value=col)
        # Écrit le nom de la colonne dans la cellule (ligne 1, colonne c)

        cell.font = font_h
        # Applique la police blanche en gras

        cell.fill = fill_h
        # Applique le fond bleu

        cell.alignment = Alignment(horizontal="center")
        # Centre le texte dans la cellule

    # -- Écriture des données (à partir de la ligne 2) --
    for r, row in enumerate(df_dict.itertuples(index=False), 2):
        # itertuples() : parcourt les lignes du DataFrame efficacement
        # index=False  : n'inclut pas l'index pandas
        # enumerate(..., 2) : commence à la ligne 2 (ligne 1 = en-têtes)

        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
            # Écrit chaque valeur dans la cellule correspondante

    # -- Largeurs des colonnes --
    widths = [20, 25, 12, 22, 18, 8, 8, 20]
    # Largeurs en caractères pour chacune des 8 colonnes :
    # variable(20), source(25), type(12), nb_uniques(22),
    # taux_manquant(18), min(8), max(8), exemple(20)

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        # get_column_letter(1) - "A", get_column_letter(2) - "B"...
        # On définit la largeur de chaque colonne pour que le fichier
        # soit lisible sans avoir à redimensionner manuellement

    wb.save(output_path)
    # Sauvegarde le fichier Excel sur le disque

    print(f" Dictionnaire sauvegardé : {output_path}  ({len(df_dict)} variables)")


# ════════════════════════════════════════════════════════════════
# FONCTION PRINCIPALE - main()
# ════════════════════════════════════════════════════════════════

def main():
    section_header("ÉTAPE 1 - DICTIONNAIRE VARIABLES MÉNAGE")

    # -- Chargement des deux fichiers sources --
    print("\n[Habitat - Section E]")
    df_habitat = read_sav(FILE_HABITAT)
    # Lit habitat_SECTION_E.sav complet (242 colonnes, 188 550 lignes)
    # On charge TOUT ici car on veut documenter TOUTES les variables

    print("\n[Individus - Section B  (pour variables CM)]")
    df_indiv = read_sav(FILE_INDIV)
    # Lit indiv_SECTION_B.sav complet (146 colonnes, 1 712 428 lignes)
    # Même logique : on veut voir toutes les variables disponibles

    # -- Construction des dictionnaires --
    dict_habitat = build_dict(df_habitat, source="habitat_SECTION_E.sav")
    # Produit un DataFrame de 242 lignes (une par variable de Section E)

    dict_indiv = build_dict(df_indiv, source="indiv_SECTION_B.sav")
    # Produit un DataFrame de 146 lignes (une par variable de Section B)

    # -- Fusion des deux dictionnaires --
    df_dict = pd.concat([dict_habitat, dict_indiv], ignore_index=True)
    # pd.concat colle les deux verticalement - 242 + 146 = 388 variables
    # ignore_index=True : recrée un index propre de 0 à 387

    print(f"\n  - {len(dict_habitat)} variables habitat  |  {len(dict_indiv)} variables individus")

    # -- Export Excel --
    output_path = OUTPUT_DIR / "dict" / "hh_dict.xlsx"
    save_dict_excel(df_dict, output_path)
    # Sauvegarde le dictionnaire de 388 variables en Excel mis en forme


if __name__ == "__main__":
    main()